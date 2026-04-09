"""Сервис шаблонов заявок и Excel-файлов для МаркЗнак и пользователей."""
import io
from datetime import datetime, timedelta, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- Внутренний формат (полный) для МаркЗнак ---

# Формат Одежда (как в расширенном шаблоне МаркЗнак / справочнике)
# Порядок колонок согласован с примером 33708 для админов.
CLOTHING_HEADERS = [
    None,  # пустой столбец, зарезервирован как в справочнике
    "GTIN",
    "Количество",
    "Тип КМ",
    "Способ выпуска",
    "Артикул",
    "Размер",
    "Дата создания",
    "Наименование товара",
    "Набор",
    "Бренд",
    "Вид товара",
    "Код ТН ВЭД",
    "Артикул",  # дублирующий артикул после кода ТН ВЭД
    "Страна производства",
    "Цвет",
    "Размер",  # дублирующий размер после цвета
    "Тип Текстиля",
    "Возраст потребителя",
    "Состав",
    "Целевой пол",
    "Статус",
    "Расширенный статус",
    "Подписан",
    "Группа ТН ВЭД",
    "Тип продукта",
    "Номер сертификата соответствия",
    "Номер свидетельства о государственной регистрации",
    "Номер декларации о соответствии",
    "Код в системе поставщика",
    "Юр. лицо",
    "Номер заказа МС",
]

# Формат Обувь (как в справочнике / шаблоне для админа)
SHOES_HEADERS = [
    "GTIN",
    "Количество",
    "Тип КМ",
    "Способ выпуска",
    "Артикул",
    "Размер",
    "Способ оплаты",
    "Идентификатор производственного заказа",
    "Дата создания",
    "Наименование товара",
    "Набор",
    "Бренд",
    "Вид обуви",
    "Код ТН ВЭД",
    "Артикул",  # дублирующий артикул
    "Страна производства",
    "Цвет",
    "Размер",
    "Состав",
    "Статус",
    "Расширенный статус",
    "Подписан",
    "Группа ТН ВЭД",
    "Материал верха",
    "Материал подкладки",
    "Материал низа / подошвы",
    "Тип продукта",
    "Номер сертификата соответствия",
    "Номер свидетельства о государственной регистрации",
    "Номер декларации о соответствии",
    "Код в системе поставщика",
    "Дата создания (техническая)",
    "Юр. лицо",
    "Номер заказа МС",
]
# Индексы (0-based) для маппинга (с учётом первого пустого столбца)
IDX_GTIN = 1
IDX_QTY = 2
IDX_ARTICLE = 5
IDX_SIZE = 6

# Стиль для столбцов, которые заполняет пользователь (подсветка)
FILL_USER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _autofit_columns(
    ws,
    *,
    min_widths: list[int] | None = None,
    max_width: int = 50,
) -> None:
    """
    Подогнать ширину колонок под содержимое с учётом минимальных ширин.
    """
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column:
            value = cell.value
            if value is None:
                continue
            # Учитываем переносы строк: берём самую длинную строку
            for part in str(value).splitlines():
                if len(part) > max_length:
                    max_length = len(part)
        base_width = 0
        if min_widths and 0 <= col_idx - 1 < len(min_widths):
            base_width = min_widths[col_idx - 1]
        # небольшой запас по бокам
        target_width = max(base_width, max_length + 2 if max_length > 0 else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            target_width, max_width
        )


def _split_shoes_composition(
    value: str | None,
) -> tuple[str, str, str, str]:
    """
    Разобрать состав обуви на отдельные поля:
    - Материал верха
    - Материал подкладки
    - Материал низа / подошвы
    Остальные части возвращаются как «чистый» состав.
    """
    if not value:
        return "", "", "", ""

    material_upper = ""
    material_lining = ""
    material_sole = ""
    other_parts: list[str] = []

    for part in value.split(";"):
        s = part.strip()
        if not s:
            continue
        lower = s.lower()
        if lower.startswith("материал верха:"):
            material_upper = s.split(":", 1)[1].strip() if ":" in s else ""
        elif lower.startswith("материал подкладки:"):
            material_lining = s.split(":", 1)[1].strip() if ":" in s else ""
        elif lower.startswith("материал низа") or lower.startswith(
            "материал низа / подошвы"
        ):
            material_sole = s.split(":", 1)[1].strip() if ":" in s else ""
        else:
            other_parts.append(s)

    clean_composition = "; ".join(other_parts)
    return clean_composition, material_upper, material_lining, material_sole


def _product_to_markznak_row(p: dict) -> list:
    """Преобразовать продукт/позицию заявки в строку расширенного шаблона МаркЗнак."""
    created = p.get("created_at")
    created_str = ""
    if created:
        if hasattr(created, "strftime"):
            created_str = created.strftime("%d.%m.%Y %H:%M")
        elif isinstance(created, str) and len(created) >= 10:
            created_str = created[:16].replace("T", " ")
    article = p.get("article") or ""
    size = p.get("size") or ""
    tnved = p.get("tnved_code") or ""
    color = p.get("color") or ""
    composition = p.get("composition") or ""
    target_gender = p.get("target_gender") or ""
    legal_entity = p.get("legal_entity") or ""
    ms_order_number = p.get("ms_order_number") or ""
    group_tnved = tnved[:4] if isinstance(tnved, str) else ""
    status = p.get("status") or "Опубликован"
    extended_status = p.get("extended_status") or status
    signed = p.get("signed") or "TRUE"

    return [
        "",  # первый пустой столбец
        p.get("gtin") or "",
        p.get("quantity") or "",
        p.get("km_type") or "Единица товара",
        p.get("release_method") or "",
        article,
        size,
        created_str,
        p.get("name") or "",
        p.get("is_set", False) and "TRUE" or "FALSE",
        p.get("brand") or "",
        p.get("category") or p.get("variant") or "",
        tnved,
        article,  # дублирующий артикул
        p.get("country") or "",
        color,
        size,  # дублирующий размер
        p.get("textile_type") or "",
        p.get("consumer_age") or "",
        composition,
        target_gender,
        status,
        extended_status,
        signed,
        group_tnved,
        p.get("product_type") or "",
        p.get("certificate_number") or "",
        p.get("witness_number") or "",
        p.get("declaration_number") or "",
        p.get("system_code") or "",
        legal_entity,
        ms_order_number,
    ]


def _product_to_markznak_row_shoes(p: dict) -> list:
    """Строка расширенного шаблона для ОБУВИ (как в обувном примере МаркЗнак)."""
    created = p.get("created_at")
    created_str = ""
    if created:
        if hasattr(created, "strftime"):
            created_str = created.strftime("%d.%m.%Y %H:%M")
        elif isinstance(created, str) and len(created) >= 10:
            created_str = created[:16].replace("T", " ")
    article = p.get("article") or ""
    size = p.get("size") or ""
    tnved = p.get("tnved_code") or ""
    raw_composition = p.get("composition") or ""
    legal_entity = p.get("legal_entity") or ""
    ms_order_number = p.get("ms_order_number") or ""
    group_tnved = tnved[:4] if isinstance(tnved, str) else ""
    status = p.get("status") or "Опубликован"
    extended_status = p.get("extended_status") or status
    signed = p.get("signed") or "TRUE"

    composition, material_upper, material_lining, material_sole = _split_shoes_composition(
        raw_composition
    )

    return [
        p.get("gtin") or "",
        p.get("quantity") or "",
        p.get("km_type") or "Единица товара",
        p.get("release_method") or "",
        article,
        size,
        p.get("payment_method") or "",
        "",  # Идентификатор производственного заказа
        created_str,
        p.get("name") or "",
        p.get("is_set", False) and "TRUE" or "FALSE",
        p.get("brand") or "",
        p.get("category") or p.get("variant") or "",
        tnved,
        article,  # дубль артикула
        p.get("country") or "",
        p.get("color") or "",
        size,
        composition,
        status,
        extended_status,
        signed,
        group_tnved,
        material_upper,
        material_lining,
        material_sole,
        p.get("product_type") or "",
        p.get("certificate_number") or "",
        p.get("witness_number") or "",
        p.get("declaration_number") or "",
        p.get("system_code") or "",
        "",  # Дата создания (техническая)
        legal_entity,
        ms_order_number,
    ]


def generate_markznak_template_excel(
    items: list[dict],
    sheet_title: str = "Заявка",
    force_mode: Optional[str] = None,
) -> bytes:
    """
    Сгенерировать расширенный шаблон Excel в формате МаркЗнак
    (все колонки CLOTHING_HEADERS).

    items — список словарей с полями: gtin, article, size, quantity, tnved_code, etc.
    """
    # Выбираем формат в зависимости от типа товара:
    # - если явно передан режим через force_mode: "shoes" / "clothing" — используем его;
    # - иначе автоопределение:
    #   обувь — ТН ВЭД начинается с 64 ИЛИ состав содержит лейблы материалов (как в обувном шаблоне);
    #   иначе одежда.
    def _is_shoes_item(p: dict) -> bool:
        tnved = p.get("tnved_code")
        if isinstance(tnved, str) and tnved.startswith("64"):
            return True
        comp = p.get("composition") or ""
        if isinstance(comp, str) and "Материал верха" in comp:
            return True
        return False

    if force_mode is not None:
        mode = force_mode.strip().lower()
        if mode == "shoes":
            is_shoes = True
        elif mode == "clothing":
            is_shoes = False
        else:
            is_shoes = any(_is_shoes_item(p) for p in items)
    else:
        is_shoes = any(_is_shoes_item(p) for p in items)
    headers = SHOES_HEADERS if is_shoes else CLOTHING_HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовки
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = border
        # Внутренний шаблон — подсветка количества не критична, но оставим её.
        if col - 1 == IDX_QTY:
            cell.fill = FILL_USER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Данные
    for row_idx, p in enumerate(items, 2):
        row_data = (
            _product_to_markznak_row_shoes(p)
            if is_shoes
            else _product_to_markznak_row(p)
        )
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # Ширина колонок: автоподгонка под содержимое c разумными минимальными значениями.
    if is_shoes:
        base_widths = [
            12,  # GTIN
            15,
            10,
            12,
            12,
            18,
            8,
            12,
            18,
            18,
            8,
            15,
            15,
            12,
            18,
            15,
            15,
            8,
            35,
            10,
            18,
            12,
            12,
            20,
            20,
            20,
            18,
            22,
            22,
            22,
            22,
            18,
            18,
            15,
        ]
    else:
        base_widths = [
            12,
            15,
            10,
            12,
            12,
            18,
            8,
            18,
            50,
            8,
            15,
            15,
            12,
            18,
            15,
            15,
            8,
            18,
            18,
            35,
            12,
            10,
            18,
            12,
            12,
            18,
            22,
            22,
            22,
            22,
            18,
            15,
        ]
    _autofit_columns(ws, min_widths=base_widths, max_width=60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_markznak_template_excel(content: bytes) -> list[dict]:
    """
    Распарсить заполненный расширенный шаблон МаркЗнак.
    Возвращает список словарей {article, size, quantity, gtin}.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result: list[dict] = []
    for row in rows:
        if not row or len(row) <= IDX_QTY:
            continue
        qty = row[IDX_QTY]
        try:
            q = int(qty) if qty is not None else 0
        except (ValueError, TypeError):
            q = 0
        if q <= 0:
            continue
        article = str(row[IDX_ARTICLE] or "").strip()
        size = str(row[IDX_SIZE] or "").strip()
        gtin = str(row[IDX_GTIN] or "").strip() if len(row) > IDX_GTIN else ""
        result.append(
            {
                "article": article,
                "size": size,
                "quantity": q,
                "gtin": gtin,
            }
        )
    return result


TZ_UTC3 = timezone(timedelta(hours=3))


def add_order_header_to_excel(
    content: bytes,
    order_number: str,
    author: str,
    created_at: datetime,
) -> bytes:
    """Добавить в начало Excel-файла заголовок заявки (Заявка №, Дата, Автор)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    ws.insert_rows(1, 4)  # Вставить 4 пустые строки сверху
    ws["A1"] = f"Заявка № {order_number}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Дата: {created_at.strftime('%d.%m.%Y %H:%M')}"
    ws["A3"] = f"Автор: {author}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# --- Пользовательские шаблоны (упрощённые) ---

# Минимальные колонки, которые видит пользователь (одежда — расширенный шаблон)
USER_TEMPLATE_HEADERS_CLOTHING = [
    "Количество",          # A
    "Артикул",             # B
    "Размер",              # C
    "Наименование товара", # D
    "Бренд",               # E
    "Вид товара",          # F
    "Код ТН ВЭД",          # G
    "Страна производства", # H
    "Цвет",                # I
    "Состав",              # J
    "Целевой пол",         # K
    "Номер заказа МС",     # L
    "Юридическое лицо",    # M
]

# Колонки для обуви (новый шаблон, без "Целевой пол" для пользователя)
USER_TEMPLATE_HEADERS_SHOES = [
    "Количество",
    "Артикул",
    "Наименование товара",
    "Бренд",
    "Вид обуви",
    "Код ТН ВЭД",
    "Страна производства",
    "Цвет",
    "Размер",
    "Материал верха",
    "Материал подкладки",
    "Материал низа / подошвы",
    "Юр. лицо",
    "Номер заказа МС",
]

USER_IDX_QTY = 0
USER_IDX_ARTICLE = 1
USER_IDX_SIZE = 2


def _extract_size_from_name(name: str | None) -> str:
    """
    Попробовать вытащить размер из наименования товара для обуви.

    Примеры:
    - "..., р. 47"  -> "47"
    - "..., размер 40" -> "40"
    - "..., р. 36-37" -> "36-37"
    Берём самую правую числовую \"группу\" вместе с возможным дефисом.
    """
    if not name:
        return ""
    s = name.strip()
    # Ищем паттерны вида "р. 47" или "размер 40"
    markers = ["р.", "р ", "размер"]
    for m in markers:
        idx = s.lower().rfind(m)
        if idx != -1:
            tail = s[idx + len(m) :].strip()
            break
    else:
        tail = s

    # Из хвоста берём последнюю "группу" с цифрами и возможным дефисом.
    tokens = tail.replace(",", " ").split()
    candidate = ""
    for tok in reversed(tokens):
        if any(ch.isdigit() for ch in tok):
            candidate = tok
            break
    return candidate.strip(",.;") if candidate else ""


def _normalize_category(category: str | None) -> str:
    """Нормализовать категорию в один из режимов шаблона."""
    if not category:
        return "clothing"
    c = category.strip().lower()
    if "обув" in c or "shoe" in c:
        return "shoes"
    return "clothing"


def generate_user_template_excel(
    products: list[dict],
    sheet_title: str = "Заявка",
    category: str | None = None,
    *,
    legal_entity: str | None = None,
    brand: str | None = None,
    country: str | None = None,
    target_gender: str | None = None,
) -> bytes:
    """
    Сгенерировать пользовательский упрощённый шаблон Excel.

    products — список словарей из справочника (повторные товары);
    если список пустой, генерируется шаблон только с заголовками (для новых товаров).
    """
    template_mode = _normalize_category(category)
    headers = (
        USER_TEMPLATE_HEADERS_SHOES
        if template_mode == "shoes"
        else USER_TEMPLATE_HEADERS_CLOTHING
    )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовки
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = border
        # Все колонки — \"пользовательские\", подсветим их
        cell.fill = FILL_USER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Предзаполненные строки для повторных товаров
    for row_idx, p in enumerate(products, 2):
        if template_mode == "shoes":
            size_val = p.get("size") or _extract_size_from_name(p.get("name"))
            (
                _,
                material_upper,
                material_lining,
                material_sole,
            ) = _split_shoes_composition(p.get("composition"))
            row_vals = [
                "",  # Количество — вводит пользователь
                p.get("article") or "",
                p.get("name") or "",
                p.get("brand") or "",
                p.get("category") or p.get("variant") or "",
                p.get("tnved_code") or "",
                p.get("country") or "",
                p.get("color") or "",
                size_val,
                material_upper,
                material_lining,
                material_sole,
                p.get("legal_entity") or "",
                "",  # Номер заказа МС заполняет пользователь
            ]
        else:
            row_vals = [
                "",  # Количество — вводит пользователь
                p.get("article") or "",
                p.get("size") or "",
                p.get("name") or "",
                p.get("brand") or "",
                p.get("category") or p.get("variant") or "",
                p.get("tnved_code") or "",
                p.get("country") or "",
                p.get("color") or "",
                p.get("composition") or "",
                p.get("target_gender") or "",
                "",  # Номер заказа МС заполняет пользователь
                p.get("legal_entity") or "",
            ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # Новый товар: одна пустая строка с предзаполненными общими полями
    if not products and any([legal_entity, brand, country, target_gender]):
        row_idx = 2
        if template_mode == "shoes":
            row_vals = [
                "",  # Количество
                "",  # Артикул
                "",  # Наименование товара
                brand or "",
                "",  # Вид обуви
                "",  # Код ТН ВЭД
                country or "",
                "",  # Цвет
                "",  # Размер
                "",  # Материал верха
                "",  # Материал подкладки
                "",  # Материал низа / подошвы
                legal_entity or "",
                "",  # Номер заказа МС
            ]
        else:
            # Одежда — расширенный шаблон с брендом/страной/полом
            row_vals = [
                "",  # Количество
                "",  # Артикул
                "",  # Размер
                "",  # Наименование товара
                brand or "",
                "",  # Вид товара
                "",  # Код ТН ВЭД
                country or "",
                "",  # Цвет
                "",  # Состав
                target_gender or "",
                "",  # Номер заказа МС
                legal_entity or "",
            ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # Ширины колонок: автоподгонка под содержимое, но не уже разумного минимума.
    if template_mode == "shoes":
        base_widths = [
            10,  # Количество
            18,  # Артикул
            40,  # Наименование товара
            18,  # Бренд
            18,  # Вид обуви
            14,  # Код ТН ВЭД
            20,  # Страна производства
            15,  # Цвет
            10,  # Размер
            25,  # Материал верха
            25,  # Материал подкладки
            25,  # Материал низа / подошвы
            18,  # Юр. лицо
            20,  # Номер заказа МС
        ]
    else:
        base_widths = [
            10,  # Количество
            18,  # Артикул
            10,  # Размер
            40,  # Наименование товара
            18,  # Бренд
            18,  # Вид товара
            14,  # Код ТН ВЭД
            20,  # Страна производства
            15,  # Цвет
            30,  # Состав
            15,  # Целевой пол
            20,  # Номер заказа МС
            20,  # Юридическое лицо
        ]

    _autofit_columns(ws, min_widths=base_widths, max_width=60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _parse_int_quantity(raw: str | int | float | None) -> int:
    """Аккуратно распарсить количество из ячейки.

    Допускаем значения вида:
    - 10
    - 10.0 / 10,0
    - "10", "10 шт", "10шт"

    Пустое значение трактуем как 0, чтобы строки без количества
    просто игнорировались и не попадали в заявку.
    """
    if raw is None or raw == "":
        return 0
    if isinstance(raw, (int, float)):
        try:
            return int(raw)
        except (ValueError, TypeError):
            return 0
    s = str(raw).strip()
    if not s:
        return 0
    # Берём только ведущую числовую часть
    num_chars: list[str] = []
    for ch in s:
        if ch.isdigit():
            num_chars.append(ch)
        elif ch in {",", "."} and num_chars:
            # допускаем разделитель, но после него не идём дальше
            break
        elif num_chars:
            # как только встретили нецифру после начала числа — останавливаемся
            break
    if not num_chars:
        return 0
    try:
        return int("".join(num_chars))
    except (ValueError, TypeError):
        return 0


def parse_user_template_excel(content: bytes) -> list[dict]:
    """
    Распарсить пользовательский упрощённый шаблон.
    Возвращает список словарей с полями как минимум:
    article, size, quantity, name, item_type, tnved_code, color, composition,
    ms_order_number, legal_entity, brand, country, target_gender.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Читаем все строки, чтобы уметь пропускать пустые/служебные строки сверху.
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        return []

    # Ищем первую непустую строку и считаем её заголовком (на случай,
    # если пользователь добавил пустые строки или какие-то надписи сверху).
    header_idx = None
    for i, row in enumerate(all_rows):
        if row and any(cell is not None and str(cell).strip() for cell in row):
            header_idx = i
            header_row = row
            break
    if header_idx is None:
        return []

    headers_lower = [
        (str(h).strip().lower() if h is not None else "") for h in header_row
    ]

    # Карта колонок для разных типов шаблонов
    clothing_header_map = {
        "количество": "qty",
        "артикул": "article",
        "размер": "size",
        "наименование товара": "name",
        "вид товара": "item_type",
        "код тн вэд": "tnved_code",
        "страна производства": "country",
        "цвет": "color",
        "состав": "composition",
        "номер заказа мс": "ms_order_number",
        "юридическое лицо": "legal_entity",
        "бренд": "brand",
        "целевой пол": "target_gender",
    }
    shoes_header_map = {
        "количество": "qty",
        "артикул": "article",
        "наименование товара": "name",
        "бренд": "brand",
        "вид обуви": "item_type",
        "код тн вэд": "tnved_code",
        "страна производства": "country",
        "цвет": "color",
        "размер": "size",
        "материал верха": "material_upper",
        "материал подкладки": "material_lining",
        "материал низа / подошвы": "material_sole",
        "целевой пол": "target_gender",
        "юр. лицо": "legal_entity",
        "номер заказа мс": "ms_order_number",
    }

    def build_index_map(header_map: dict[str, str]) -> dict[str, int]:
        idx: dict[str, int] = {}
        for i, h in enumerate(headers_lower):
            key = h.strip().lower()
            if key in header_map:
                idx[header_map[key]] = i
        return idx

    shoes_idx = build_index_map(shoes_header_map)
    clothing_idx = build_index_map(clothing_header_map)

    # Определяем тип шаблона:
    # - обувь: есть хотя бы одна из колонок материалов ИЛИ явно «Вид обуви» без «Вид товара»;
    # - иначе считаем, что это одежда.
    has_shoes_materials = any(
        k in shoes_idx for k in ("material_upper", "material_lining", "material_sole")
    )
    has_shoes_item_type = "item_type" in shoes_idx and "item_type" not in clothing_idx

    if has_shoes_materials or has_shoes_item_type:
        idx = shoes_idx
        template_mode = "shoes"
    else:
        idx = clothing_idx
        template_mode = "clothing"

    # На случай, если колонка количества называется чуть иначе, чем ожидается,
    # пробуем подобрать её по подстроке «колич».
    if "qty" not in idx:
        for i, h in enumerate(headers_lower):
            if "колич" in h:
                idx["qty"] = i
                break

    def get_cell(row: tuple, key: str) -> str:
        col_idx = idx.get(key)
        if col_idx is None or col_idx >= len(row):
            return ""
        val = row[col_idx]
        return str(val).strip() if val is not None else ""

    # Последующие строки после заголовка — это данные.
    rows_iter = iter(all_rows[header_idx + 1 :])

    result: list[dict] = []
    errors: list[str] = []
    skipped_qty_zero = 0
    skipped_missing_keys = 0
    row_number = header_idx + 1  # номер строки заголовка в файле
    for row in rows_iter:
        row_number += 1
        if not row:
            continue

        qty_raw = get_cell(row, "qty")
        qty = _parse_int_quantity(qty_raw)
        if qty <= 0:
            skipped_qty_zero += 1
            continue

        article = get_cell(row, "article")
        size = get_cell(row, "size")

        name = get_cell(row, "name")
        item_type = get_cell(row, "item_type")
        tnved = get_cell(row, "tnved_code")
        color = get_cell(row, "color")
        ms_order_number = get_cell(row, "ms_order_number")
        legal_entity = get_cell(row, "legal_entity")
        brand = get_cell(row, "brand") if "brand" in idx else ""
        country = get_cell(row, "country") if "country" in idx else ""
        target_gender = (
            get_cell(row, "target_gender") if "target_gender" in idx else ""
        )

        if template_mode == "shoes":
            # Для обуви материалы вводятся отдельными колонками и сохраняются
            # в составе, чтобы затем разложить их по полям админского шаблона.
            material_upper = get_cell(row, "material_upper")
            material_lining = get_cell(row, "material_lining")
            material_sole = get_cell(row, "material_sole")
            parts: list[str] = []
            if material_upper:
                parts.append(f"Материал верха: {material_upper}")
            if material_lining:
                parts.append(f"Материал подкладки: {material_lining}")
            if material_sole:
                parts.append(f"Материал низа / подошвы: {material_sole}")
            composition = "; ".join(parts)
        else:
            composition = get_cell(row, "composition")

        # Для одежды проверяем обязательные поля; для обуви сохраняем
        # текущую гибкость (требуем только количество > 0, article и size).
        # Для новых шаблонов логика единая:
        # - для строк с qty > 0 все поля, кроме "Номер заказа МС", должны быть заполнены;
        # - проверяем только те колонки, которые реально присутствуют в файле (idx).
        missing: list[str] = []
        if template_mode == "clothing":
            required_fields: list[tuple[str, str]] = [
                ("article", "Артикул"),
                ("size", "Размер"),
                ("name", "Наименование товара"),
                ("item_type", "Вид товара"),
                ("tnved_code", "Код ТН ВЭД"),
                ("color", "Цвет"),
                ("composition", "Состав"),
                ("brand", "Бренд"),
                ("target_gender", "Целевой пол"),
                ("legal_entity", "Юридическое лицо"),
            ]
            for key, label in required_fields:
                if key not in idx:
                    continue
                if key == "article":
                    value = article
                elif key == "size":
                    value = size
                elif key == "name":
                    value = name
                elif key == "item_type":
                    value = item_type
                elif key == "tnved_code":
                    value = tnved
                elif key == "country":
                    value = country
                elif key == "color":
                    value = color
                elif key == "composition":
                    value = composition
                elif key == "brand":
                    value = brand
                elif key == "target_gender":
                    value = target_gender
                elif key == "legal_entity":
                    value = legal_entity
                else:
                    value = ""
                if not value:
                    missing.append(label)
        else:
            required_fields_shoes: list[tuple[str, str]] = [
                ("article", "Артикул"),
                ("name", "Наименование товара"),
                ("brand", "Бренд"),
                ("item_type", "Вид обуви"),
                ("tnved_code", "Код ТН ВЭД"),
                ("color", "Цвет"),
                ("size", "Размер"),
                ("material_upper", "Материал верха"),
                ("material_lining", "Материал подкладки"),
                ("material_sole", "Материал низа / подошвы"),
                ("legal_entity", "Юр. лицо"),
            ]
            for key, label in required_fields_shoes:
                if key not in idx:
                    continue
                if key == "article":
                    value = article
                elif key == "name":
                    value = name
                elif key == "brand":
                    value = brand
                elif key == "item_type":
                    value = item_type
                elif key == "tnved_code":
                    value = tnved
                elif key == "country":
                    value = country
                elif key == "color":
                    value = color
                elif key == "size":
                    value = size
                elif key == "material_upper":
                    value = material_upper
                elif key == "material_lining":
                    value = material_lining
                elif key == "material_sole":
                    value = material_sole
                elif key == "target_gender":
                    value = target_gender
                elif key == "legal_entity":
                    value = legal_entity
                else:
                    value = ""
                if not value:
                    missing.append(label)

        if missing:
            errors.append(
                f"Строка {row_number}: не заполнены обязательные поля: {', '.join(missing)}"
            )
            # пропускаем строку, но продолжаем разбор файла
            continue

        result.append(
            {
                "article": article,
                "size": size,
                "quantity": qty,
                "name": name,
                "item_type": item_type,
                "tnved_code": tnved,
                "color": color,
                "composition": composition,
                "ms_order_number": ms_order_number,
                "legal_entity": legal_entity,
                "brand": brand or None,
                "country": country or None,
                "target_gender": target_gender or None,
            }
        )

    if not result:
        # Если ни одной валидной позиции не получилось, но при этом есть
        # подробные ошибки по незаполненным полям — возвращаем именно их,
        # чтобы пользователь увидел, какие колонки нужно заполнить.
        if errors:
            raise ValueError("; ".join(errors))

        # Иначе формируем агрегированное сообщение по количеству/ключам.
        parts: list[str] = []
        if skipped_qty_zero:
            parts.append(
                f"строк с пустым или нулевым количеством: {skipped_qty_zero}"
            )
        if skipped_missing_keys:
            parts.append(
                f"строк без артикула или размера: {skipped_missing_keys}"
            )
        details = "; ".join(parts) if parts else "все строки пустые или нераспознанные"
        raise ValueError(
            f"В шаблоне нет корректных позиций (Количество > 0). Диагностика: {details}"
        )

    if errors:
        # Есть валидные строки, но часть строк с ошибками — тоже покажем детали.
        raise ValueError("; ".join(errors))

    return result
