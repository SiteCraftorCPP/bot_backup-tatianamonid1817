"""Excel file generation for orders."""
import io
import re
from datetime import datetime, timedelta, timezone
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TZ_UTC3 = timezone(timedelta(hours=3))


def generate_order_excel(
    order_number: str,
    author_username: str | None,
    author_full_name: str | None,
    created_at: datetime,
    items: list[dict],
) -> bytes:
    """Generate Excel file for order. Returns file bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"
    
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Заявка № {order_number}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    row = 2
    ws[f"A{row}"] = "Дата:"
    ws[f"B{row}"] = created_at.astimezone(TZ_UTC3).strftime("%d.%m.%Y %H:%M")
    # Пропускаем строку 3, данные «Автор» больше не выводим
    row += 2
    
    # Table header
    headers = ["№", "Артикул", "Наименование", "Цвет", "Размер", "Кол-во", "Код ТН ВЭД", "Состав"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1
    
    for i, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=item.get("article", "") or "").border = border
        ws.cell(row=row, column=3, value=item.get("name", "") or "").border = border
        ws.cell(row=row, column=4, value=item.get("color", "") or "").border = border
        ws.cell(row=row, column=5, value=item.get("size", "") or "").border = border
        ws.cell(row=row, column=6, value=item.get("quantity", 1)).border = border
        ws.cell(row=row, column=7, value=item.get("tnved_code", "") or "").border = border
        ws.cell(row=row, column=8, value=item.get("composition", "") or "").border = border
        row += 1
    
    # Column widths
    for col, width in enumerate([6, 15, 40, 15, 10, 8, 15, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _safe_download_filename_fragment(order_number: str) -> str:
    """Фрагмент имени файла без символов, запрещённых в путях Windows."""
    if not order_number:
        return "order"
    t = str(order_number).strip()
    for bad in '<>:"/\\|?*\n\r\t':
        t = t.replace(bad, "_")
    t = re.sub(r"\s+", "_", t)
    return t or "order"


def get_order_excel_download_filename(order_number: str) -> str:
    """Имя скачиваемого Excel по заявке (совпадает с номером в системе)."""
    return f"Заявка_{_safe_download_filename_fragment(order_number)}.xlsx"


def get_markznak_download_filename(order_number: str) -> str:
    """Имя файла шаблона Маркзнак для заявки."""
    return f"Заявка_{_safe_download_filename_fragment(order_number)}_markznak.xlsx"


def content_disposition_attachment(filename: str) -> str:
    """Заголовок Content-Disposition: кириллица через RFC 5987 (значение — только ASCII)."""
    ascii_name = "".join(
        c if 32 <= ord(c) < 127 and c not in '"\\' else "_"
        for c in filename
    )
    ascii_name = re.sub(r"_+", "_", ascii_name).strip("_") or "download.xlsx"
    utf8_part = quote(filename, safe="")
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_part}"
