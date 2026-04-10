"""Tests for orders API: create orders and download Excel."""
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Product, Order, OrderItem, User
from tests.conftest import FakeSheetsService


@pytest.mark.asyncio
async def test_create_order_with_items(client: AsyncClient, test_db_session: AsyncSession, fake_sheets: FakeSheetsService):
    # Prepare products in DB
    p1 = Product(
        article="33708white",
        name="Джемпер женский, арт. 33708white",
        brand="Brand",
        color="white",
        tnved_code="6202400009",
        composition="100% хлопок",
        country="Россия",
        target_gender="ЖЕНСКИЙ",
        category="ДЖЕМПЕР",
        legal_entity="Акс Кэпитал",
        size="S",
    )
    p2 = Product(
        article="33708black",
        name="Джемпер женский, арт. 33708black",
        brand="Brand",
        color="black",
        tnved_code="6202400009",
        composition="100% хлопок",
        country="Россия",
        target_gender="ЖЕНСКИЙ",
        category="ДЖЕМПЕР",
        legal_entity="Акс Кэпитал",
        size="M",
    )
    test_db_session.add_all([p1, p2])
    await test_db_session.flush()

    payload = {
        "author_telegram_id": 111,
        "author_username": "tester",
        "author_full_name": "Test User",
        "order_type": "Ламода",
        "ms_order_number": "MS-123",
        "comment": "Тестовая заявка",
        "items": [
            {"product_id": p1.id, "size": "S", "quantity": 2},
            {"product_id": p2.id, "size": "M", "quantity": 3},
        ],
    }

    resp = await client.post("/orders/", json=payload)
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "создана"
    assert len(data["items"]) == 2

    # Sheets registry should receive one append call with basic info
    assert len(fake_sheets.append_calls) == 1
    order_number, created_at, author, items_summary, status, *_ = fake_sheets.append_calls[0]
    assert "_" in order_number and order_number.split("_", 1)[0].isdigit()
    # Публичный номер: id + числовая часть артикула (33708 из 33708white)
    oid, art_part = order_number.split("_", 1)
    assert art_part == "33708"
    # items_summary should mention both positions by quantity
    assert "x2" in items_summary and "x3" in items_summary
    assert status == "создана"


@pytest.mark.asyncio
async def test_create_order_number_uses_article_from_name_when_article_is_text(
    client: AsyncClient,
    fake_sheets: FakeSheetsService,
):
    payload = {
        "author_telegram_id": 333,
        "author_username": "tester_text_article",
        "author_full_name": "Tester Text Article",
        "order_type": "Ламода",
        "items": [
            {
                "size": "2XL",
                "quantity": 4,
                "article": "Джемпер",
                "name": "Джемпер мужской, арт. 33708whitebrown, размер 2XL",
                "brand": "FLOW LAB",
            }
        ],
    }
    resp = await client.post("/orders/", json=payload)
    assert resp.status_code == 200
    data = resp.json()
    assert "_" in data["number"]
    _, art_part = data["number"].split("_", 1)
    assert art_part == "33708"
    # Проверяем также номер, ушедший в Google Sheets
    order_number, *_ = fake_sheets.append_calls[-1]
    _, art_part_sheet = order_number.split("_", 1)
    assert art_part_sheet == "33708"


@pytest.mark.asyncio
async def test_download_order_excel_structure(client: AsyncClient, test_db_session: AsyncSession):
    # Create user, order and items directly in DB
    user = User(telegram_id=999, username="tester", full_name="Tester")
    test_db_session.add(user)
    await test_db_session.flush()

    order = Order(
        number="2026-02-999",
        author_id=user.id,
        status="создана",
        order_type="Ламода",
    )
    test_db_session.add(order)
    await test_db_session.flush()

    item = OrderItem(
        order_id=order.id,
        product_id=None,
        size="L",
        quantity=5,
        article="33708white",
        name="Джемпер женский, арт. 33708white",
        color="white",
        tnved_code="6202400009",
        composition="100% хлопок",
    )
    test_db_session.add(item)
    await test_db_session.commit()

    resp = await client.get(f"/orders/{order.id}/excel")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    # Header row
    assert ws["A1"].value.startswith("Заявка № ")
    # Table headers (row 4)
    headers = [ws.cell(row=4, column=col).value for col in range(1, 9)]
    assert headers[:4] == ["№", "Артикул", "Наименование", "Цвет"]

    # First data row (row 5)
    art = ws["B5"].value
    qty = ws["F5"].value
    assert art == "33708white"
    assert qty == 5


@pytest.mark.asyncio
async def test_order_attachment_register_and_get(client: AsyncClient, test_db_session: AsyncSession):
    user = User(telegram_id=222, username="u2", full_name="User Two")
    test_db_session.add(user)
    await test_db_session.flush()
    order = Order(number="w-testattach", author_id=user.id, status="создана")
    test_db_session.add(order)
    await test_db_session.commit()

    r403 = await client.post(
        f"/orders/{order.id}/attachments",
        params={"author_telegram_id": 999},
        json={"telegram_file_id": "AgACAgIAAxkBAAIB", "file_name": "a.xlsx"},
    )
    assert r403.status_code == 403

    r_ok = await client.post(
        f"/orders/{order.id}/attachments",
        params={"author_telegram_id": 222},
        json={"telegram_file_id": "AgACAgIAAxkBAAIB", "file_name": "3215.xlsx"},
    )
    assert r_ok.status_code == 200
    att = r_ok.json()
    assert att["file_name"] == "3215.xlsx"
    assert att["telegram_file_id"] == "AgACAgIAAxkBAAIB"

    r_get = await client.get(f"/orders/{order.id}")
    assert r_get.status_code == 200
    body = r_get.json()
    assert len(body.get("extra_attachments") or []) == 1
    assert body["extra_attachments"][0]["file_name"] == "3215.xlsx"


@pytest.mark.asyncio
async def test_repair_responsible_telegram(client: AsyncClient, test_db_session: AsyncSession):
    author = User(telegram_id=100, username="auth", role="user")
    admin = User(telegram_id=555_555, username="julia_test", role="admin")
    test_db_session.add_all([author, admin])
    await test_db_session.flush()
    order = Order(
        number="w-repair-tg",
        author_id=author.id,
        status="в работе",
        responsible_telegram_id=999_999,
        responsible_username="@julia_test",
    )
    test_db_session.add(order)
    await test_db_session.commit()

    r_forbid = await client.post(
        "/orders/repair-responsible-telegram",
        params={"telegram_id": 555_555, "requester_telegram_id": 111},
    )
    assert r_forbid.status_code == 403

    r = await client.post(
        "/orders/repair-responsible-telegram",
        params={"telegram_id": 555_555, "requester_telegram_id": 555_555},
    )
    assert r.status_code == 200
    assert r.json()["fixed"] == 1

    r2 = await client.get(f"/orders/{order.id}")
    assert r2.status_code == 200
    assert r2.json()["responsible_telegram_id"] == 555_555


@pytest.mark.asyncio
async def test_unassign_responsible_by_telegram(client: AsyncClient, test_db_session: AsyncSession):
    adm = User(telegram_id=50, username="boss", role="admin")
    auth = User(telegram_id=200, username="cli", role="user")
    test_db_session.add_all([adm, auth])
    await test_db_session.flush()
    o = Order(
        number="w-unas-tg",
        author_id=auth.id,
        status="в работе",
        responsible_telegram_id=77,
        responsible_username="gone",
    )
    test_db_session.add(o)
    await test_db_session.commit()

    r403 = await client.post(
        "/orders/unassign-responsible-by-telegram",
        params={"target_telegram_id": 77, "requester_telegram_id": 999},
    )
    assert r403.status_code == 403

    r = await client.post(
        "/orders/unassign-responsible-by-telegram",
        params={"target_telegram_id": 77, "requester_telegram_id": 50},
    )
    assert r.status_code == 200
    assert r.json()["unassigned"] == 1
    r2 = await client.get(f"/orders/{o.id}")
    assert r2.json()["responsible_telegram_id"] is None
    assert r2.json().get("responsible_username") in (None, "")
