"""Tests for template-related API endpoints."""
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Product


@pytest.mark.asyncio
async def test_get_template_by_article_filters_products(client: AsyncClient, test_db_session: AsyncSession):
    # Prepare products for two different articles
    products = [
        Product(
            article="33708white",
            name="Джемпер женский, арт. 33708white, размер S",
            brand="Brand",
            color="white",
            tnved_code="6202400009",
            composition="100% хлопок",
            country="Россия",
            target_gender="ЖЕНСКИЙ",
            category="ДЖЕМПЕР",
            legal_entity="Акс Кэпитал",
            size="S",
        ),
        Product(
            article="33708black",
            name="Джемпер женский, арт. 33708black, размер M",
            brand="Brand",
            color="black",
            tnved_code="6202400009",
            composition="100% хлопок",
            country="Россия",
            target_gender="ЖЕНСКИЙ",
            category="ДЖЕМПЕР",
            legal_entity="Акс Кэпитал",
            size="M",
        ),
        Product(
            article="042purple",
            name="Куртка женская, арт. 042purple, размер L",
            brand="Brand",
            color="purple",
            tnved_code="6202400009",
            composition="100% полиэстер",
            country="Киргизия",
            target_gender="ЖЕНСКИЙ",
            category="КУРТКА",
            legal_entity="Акс Кэпитал",
            size="L",
        ),
    ]
    test_db_session.add_all(products)
    await test_db_session.commit()

    resp = await client.get("/products/template", params={"article": "33708"})
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    # Header row of user template
    headers = [ws.cell(row=1, column=col).value for col in range(1, 10)]
    assert headers[:3] == ["Количество", "Артикул", "Размер"]

    # Data rows should only contain 33708* articles in user template
    articles = [ws.cell(row=row, column=2).value for row in range(2, ws.max_row + 1)]
    assert set(articles) == {"33708white", "33708black"}

    # Quantity column should be empty for all rows
    qty_values = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
    assert all(v in (None, "", 0) for v in qty_values)


@pytest.mark.asyncio
async def test_create_order_from_template_happy_path(client: AsyncClient, test_db_session: AsyncSession):
    # Prepare products for one article with several sizes
    base_kwargs = dict(
        name="Джинсы женские, арт. 8069white",
        brand="Bronks",
        color="white",
        tnved_code="6204623100",
        composition="75% хлопок, 15% вискоза, 10% полиэстер",
        country="Киргизия",
        target_gender="ЖЕНСКИЙ",
        category="ДЖИНСЫ",
        legal_entity="Акс Кэпитал",
    )
    products = [
        Product(article="8069white", size="S", **base_kwargs),
        Product(article="8069white", size="M", **base_kwargs),
        Product(article="8069white", size="L", **base_kwargs),
    ]
    test_db_session.add_all(products)
    await test_db_session.commit()

    # Get user template from API
    resp = await client.get("/products/template", params={"article": "8069white"})
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    # Fill quantities for first two rows in user template
    ws.cell(row=2, column=1, value=2)  # Количество for size S
    ws.cell(row=3, column=1, value=3)  # Количество for size M

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {"file": ("template.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {
        "author_telegram_id": "222",
        "author_username": "tester2",
        "author_full_name": "Tester Two",
        "order_type": "Ламода",
        "ms_order_number": "MS-TEMPLATE-1",
        "comment": "Заявка по шаблону",
    }

    resp2 = await client.post("/orders/from_template", data=data, files=files)
    assert resp2.status_code == 200
    order = resp2.json()

    assert len(order["items"]) == 2
    quantities = sorted(i["quantity"] for i in order["items"])
    assert quantities == [2, 3]


@pytest.mark.asyncio
async def test_create_order_from_template_persists_ms_in_markznak(
    client: AsyncClient, test_db_session: AsyncSession,
):
    """Номер заказа МС из колонки пользовательского шаблона попадает в Excel МаркЗнак для админа."""
    base_kwargs = dict(
        name="Тест, арт. MS-ART-1",
        brand="Brand",
        color="red",
        tnved_code="6204623100",
        composition="100% хлопок",
        country="Киргизия",
        target_gender="ЖЕНСКИЙ",
        category="ДЖИНСЫ",
        legal_entity="Акс Кэпитал",
    )
    test_db_session.add(Product(article="MS-ART-1", size="M", **base_kwargs))
    await test_db_session.commit()

    resp = await client.get("/products/template", params={"article": "MS-ART-1"})
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Колонка L — «Номер заказа МС» в одежде
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=12, value="MS-LINE-999")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {
        "file": (
            "template.xlsx",
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    data = {
        "author_telegram_id": "333",
        "author_username": "msuser",
        "author_full_name": "MS Test",
    }

    resp2 = await client.post("/orders/from_template", data=data, files=files)
    assert resp2.status_code == 200, resp2.text
    order = resp2.json()
    assert order["items"][0].get("ms_order_number") == "MS-LINE-999"
    assert order.get("ms_order_number") == "MS-LINE-999"

    oid = order["id"]
    r3 = await client.get(f"/orders/{oid}/markznak_excel")
    assert r3.status_code == 200
    wb2 = load_workbook(io.BytesIO(r3.content))
    ws2 = wb2.active
    headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    col_ms = next(i + 1 for i, h in enumerate(headers) if h == "Номер заказа МС")
    assert ws2.cell(row=2, column=col_ms).value == "MS-LINE-999"


@pytest.mark.asyncio
async def test_create_order_from_template_respects_brand_and_legal_entity(
    client: AsyncClient,
    test_db_session: AsyncSession,
):
    """Не смешивать дубли article+size с разными брендом/ЮЛ."""
    p_line17 = Product(
        article="2705black",
        size="30",
        name="Брюки женские, арт. 2705black, размер 30",
        brand="Line 17",
        color="Черный",
        tnved_code="6203429000",
        composition="97% хлопок, 3% эластан",
        country="Россия",
        target_gender="ЖЕНСКИЙ",
        category="Брюки",
        legal_entity="Малец",
    )
    p_flowlab = Product(
        article="2705black",
        size="30",
        name="Брюки женские, арт. 2705black, размер 30",
        brand="Flow Lab",
        color="Черный",
        tnved_code="6203429000",
        composition="97% хлопок, 3% эластан",
        country="Россия",
        target_gender="ЖЕНСКИЙ",
        category="Брюки",
        legal_entity="Чайковский",
    )
    test_db_session.add_all([p_line17, p_flowlab])
    await test_db_session.commit()
    await test_db_session.refresh(p_line17)
    await test_db_session.refresh(p_flowlab)

    resp = await client.get("/products/template", params={"article": "2705black"})
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    # Находим строку именно для Flow Lab + Чайковский и задаём количество.
    target_row = None
    for row in range(2, ws.max_row + 1):
        article = ws.cell(row=row, column=2).value
        size = ws.cell(row=row, column=3).value
        brand = ws.cell(row=row, column=5).value
        legal_entity = ws.cell(row=row, column=13).value
        if (
            str(article or "").strip() == "2705black"
            and str(size or "").strip() == "30"
            and str(brand or "").strip().lower() == "flow lab"
            and str(legal_entity or "").strip().lower() == "чайковский"
        ):
            target_row = row
            break

    assert target_row is not None, "В шаблоне не найдена строка Flow Lab + Чайковский"
    ws.cell(row=target_row, column=1, value=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {
        "file": (
            "template_2705.xlsx",
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    data = {
        "author_telegram_id": "777",
        "author_username": "tester_critical",
        "author_full_name": "Tester Critical",
    }

    resp2 = await client.post("/orders/from_template", data=data, files=files)
    assert resp2.status_code == 200
    order = resp2.json()
    assert len(order["items"]) == 1

    item = order["items"][0]
    assert item["product_id"] == p_flowlab.id
    assert item["brand"] == "Flow Lab"
    assert item["legal_entity"] == "Чайковский"


@pytest.mark.asyncio
async def test_template_legal_entities_and_countries_and_filters(
    client: AsyncClient,
    test_db_session: AsyncSession,
):
    """Повторный товар: списки ЮЛ/стран и сужение шаблона по query-параметрам."""
    p1 = Product(
        article="2801",
        size="M",
        name="Товар 2801 Малец КНР",
        brand="B",
        color="black",
        tnved_code="6202400009",
        composition="cotton",
        country="КНР",
        target_gender="ЖЕНСКИЙ",
        category="X",
        legal_entity="Малец",
    )
    p2 = Product(
        article="2801",
        size="M",
        name="Товар 2801 Банишевский РФ",
        brand="B",
        color="black",
        tnved_code="6202400009",
        composition="cotton",
        country="Россия",
        target_gender="ЖЕНСКИЙ",
        category="X",
        legal_entity="Банишевский",
    )
    p3 = Product(
        article="2801",
        size="L",
        name="Товар 2801 Малец Киргизия",
        brand="B",
        color="black",
        tnved_code="6202400009",
        composition="cotton",
        country="Киргизия",
        target_gender="ЖЕНСКИЙ",
        category="X",
        legal_entity="Малец",
    )
    test_db_session.add_all([p1, p2, p3])
    await test_db_session.commit()

    r_le = await client.get("/products/template/legal_entities", params={"article": "2801"})
    assert r_le.status_code == 200
    les = sorted(r_le.json())
    assert les == ["Банишевский", "Малец"]

    r_cn = await client.get(
        "/products/template/countries",
        params={"article": "2801", "legal_entity": "Малец"},
    )
    assert r_cn.status_code == 200
    countries = sorted(r_cn.json())
    assert countries == ["КНР", "Киргизия"]

    r_tpl = await client.get(
        "/products/template",
        params={"article": "2801", "legal_entity": "Малец", "country": "КНР"},
    )
    assert r_tpl.status_code == 200
    wb = load_workbook(io.BytesIO(r_tpl.content))
    ws = wb.active
    articles = [ws.cell(row=row, column=2).value for row in range(2, ws.max_row + 1)]
    assert articles == ["2801"]
    sizes = [ws.cell(row=row, column=3).value for row in range(2, ws.max_row + 1)]
    assert sizes == ["M"]


@pytest.mark.asyncio
async def test_template_countries_includes_rows_when_legal_entity_has_trailing_space(
    client: AsyncClient,
    test_db_session: AsyncSession,
):
    """ЮЛ с пробелом в конце в БД не должен выпадать из фильтра — иначе «теряется» страна."""
    p_ru = Product(
        article="spacetest",
        size="S",
        name="Space test RU",
        brand="B",
        legal_entity="Банишевский ",  # как часто бывает в выгрузках
        country="Россия",
    )
    p_kg = Product(
        article="spacetest",
        size="M",
        name="Space test KG",
        brand="B",
        legal_entity="Банишевский",
        country="Киргизия",
    )
    test_db_session.add_all([p_ru, p_kg])
    await test_db_session.commit()

    r_le = await client.get("/products/template/legal_entities", params={"article": "spacetest"})
    assert r_le.status_code == 200
    assert r_le.json() == ["Банишевский"]

    r_cn = await client.get(
        "/products/template/countries",
        params={"article": "spacetest", "legal_entity": "Банишевский"},
    )
    assert r_cn.status_code == 200
    assert sorted(r_cn.json()) == ["Киргизия", "Россия"]


@pytest.mark.asyncio
async def test_template_repeat_flow_is_category_scoped(
    client: AsyncClient,
    test_db_session: AsyncSession,
):
    """Категория должна жестко ограничивать ЮЛ/страны/шаблон своим справочником."""
    p_clothing = Product(
        article="catmix",
        size="S",
        name="Одежда КНР",
        brand="B",
        legal_entity="ЮЛ1",
        country="КНР",
        tnved_code="6202400009",
    )
    p_shoes = Product(
        article="catmix",
        size="40",
        name="Обувь РФ",
        brand="B",
        legal_entity="ЮЛ2",
        country="Россия",
        tnved_code="6403990000",
    )
    test_db_session.add_all([p_clothing, p_shoes])
    await test_db_session.commit()

    r_le_clothes = await client.get(
        "/products/template/legal_entities",
        params={"article": "catmix", "category": "Одежда"},
    )
    assert r_le_clothes.status_code == 200
    assert r_le_clothes.json() == ["ЮЛ1"]

    r_le_shoes = await client.get(
        "/products/template/legal_entities",
        params={"article": "catmix", "category": "Обувь"},
    )
    assert r_le_shoes.status_code == 200
    assert r_le_shoes.json() == ["ЮЛ2"]

    r_cn_clothes = await client.get(
        "/products/template/countries",
        params={"article": "catmix", "category": "Одежда", "legal_entity": "ЮЛ1"},
    )
    assert r_cn_clothes.status_code == 200
    assert r_cn_clothes.json() == ["КНР"]

    r_cn_shoes = await client.get(
        "/products/template/countries",
        params={"article": "catmix", "category": "Обувь", "legal_entity": "ЮЛ2"},
    )
    assert r_cn_shoes.status_code == 200
    assert r_cn_shoes.json() == ["Россия"]

    r_tpl_clothes = await client.get(
        "/products/template",
        params={"article": "catmix", "category": "Одежда"},
    )
    assert r_tpl_clothes.status_code == 200
    wb_c = load_workbook(io.BytesIO(r_tpl_clothes.content))
    ws_c = wb_c.active
    arts_c = [ws_c.cell(row=r, column=2).value for r in range(2, ws_c.max_row + 1)]
    assert arts_c == ["catmix"]
    sizes_c = [str(ws_c.cell(row=r, column=3).value) for r in range(2, ws_c.max_row + 1)]
    assert sizes_c == ["S"]

    r_tpl_shoes = await client.get(
        "/products/template",
        params={"article": "catmix", "category": "Обувь"},
    )
    assert r_tpl_shoes.status_code == 200
    wb_s = load_workbook(io.BytesIO(r_tpl_shoes.content))
    ws_s = wb_s.active
    sizes_s = [str(ws_s.cell(row=r, column=9).value) for r in range(2, ws_s.max_row + 1)]
    assert sizes_s == ["40"]
